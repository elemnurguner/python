from openpyxl import load_workbook

#exel dosyasını yükle
dosya_adi="ogrenciler.xlsx"
workbook=load_workbook(dosya_adi)
sheet=workbook.active

#yeni bir  öğrenci ekle
sheet.append(["Nur", "Yılmaz", 24, 85])
#mevcut bir öğrencinin notunu güncelle
for row in sheet.iter_rows(min_row=2,max_col=4,max_row=sheet.max_row):
    if row[0].value=="Ali":
        row[3].value=90

#dosyayı güncellenmiş haliyle kaydet
workbook.save(dosya_adi)
print(f"{dosya_adi} adlı dosya güncellendi.")