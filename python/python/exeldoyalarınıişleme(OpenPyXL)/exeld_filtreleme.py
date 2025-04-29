from openpyxl import load_workbook
#exel dosyasını yükle
dosya_adi="ogrenciler.xlsx"
workbook=load_workbook(dosya_adi)
sheet=workbook.active

#notu 90'dan büyük olan öğrencileri bul
for  row in sheet.iter_rows(min_row=2,values_only=True):
    ad,soyad,yas,notu=row
    if notu>90:
        print(f"{ad} {soyad} adlı öğrencinin notu {notu} oldugundan  90'dan büyük.")