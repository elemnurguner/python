from openpyxl import load_workbook  

#exel dosyasını yükle
dosya_adi="ogrenciler.xlsx"
workbook=load_workbook(dosya_adi)
sheet=workbook.active

#exel dosyasındaki verileri oku ve yazdır 
for row in sheet.iter_rows(values_only=True):#baslık satırını atla 
    ad, soyad, yas, notu=row
    print(f"Ad:{ad},Soyad: {soyad},Yaş: {yas},Not: {notu}")